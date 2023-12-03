import argparse
import openpyxl
from openpyxl.styles import Font
import re


def move_column(sheet, source_col_idx, target_col_idx):

    # Read data from source column
    source_data = [sheet.cell(row=row, column=source_col_idx).value for row in range(1, sheet.max_row + 1)]

    # Write data to target column
    for row, value in enumerate(source_data, start=1):
        sheet.cell(row=row, column=target_col_idx).value = value
    
    # Clear the data in source_col_idx
    for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=source_col_idx).value = None
    

def modify_stmt(fp):
    # Load the spreadsheet
    wb = openpyxl.load_workbook(fp)
    ws = wb.active

    # Change the font and number format to match the Credits (Calibri, 2 decimals with “,”)
    font = Font(name='Calibri')
    num_format = '0.00'
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.number_format = num_format

    # Change the date format to MM/DD/YY
    date_format = 'MM/DD/YY'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = date_format

    # Delete row 8
    ws.delete_rows(8)

    # Clear column D
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.value = ''

    # Move column B to column F and column C to column B
    move_column(ws, 2, 6)
    move_column(ws, 3, 2)

    # Create new sheets 'Credits' and 'Debits'
    credits_sheet = wb.create_sheet('Credits')
    debits_sheet = wb.create_sheet('Debits')

    # Copy rows to 'Credits' and 'Debits' sheets
    for row in wb.active.iter_rows(min_row=8):
        if row[1].value > 0:
            credits_sheet.append((cell.value for cell in row))
        elif row[1].value <= 0:
            debits_sheet.append((cell.value for cell in row))

    # Make the date and number format of the 'Credits' and 'Debits' sheets the same as the active sheet
    for sheet in [credits_sheet, debits_sheet]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font
                cell.number_format = num_format
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = date_format

    # Save the modified spreadsheet
    wb.save('stmt_1.xlsx')
    return wb


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("filepath", help="The filepath of the stmt.xlsx spreadsheeet")
    args = parser.parse_args()

    modify_stmt(args.filepath)
